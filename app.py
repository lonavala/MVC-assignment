from flask import Flask, request, render_template
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from werkzeug.utils import secure_filename
import pandas as pd

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = 'mysql+pymysql://lonavala:shivasai#123@localhost/priyanka'

db = SQLAlchemy(app)


# create table using sqlalchemy
class Data(db.Model):
    id = db.Column('id', db.Integer(), primary_key=True)
    cpart = db.Column('Child Part Number', db.String(20))
    cdesc = db.Column('Child Part Description', db.String(50))
    itemref = db.Column('item reference number', db.Integer)
    qty = db.Column('quantity production', db.Integer)


db.create_all()


@app.route('/home', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        '''Accepting file from frond end'''
        f = request.files['filename']  # to get file from front
        f.save(secure_filename(f.filename))  # to save file in local storage
        wb = openpyxl.load_workbook(f.filename)   # open that file using openpyxl
        sheet = wb.active

        '''Adding data into database'''
        try:
            for i in range(2, sheet.max_row):   # for loop till maximum row
                part = sheet.cell(row=i, column=1).value
                desc = sheet.cell(row=i, column=2).value
                item = sheet.cell(row=i, column=3).value
                qty = sheet.cell(row=i, column=4).value
                data = Data(cpart=part, cdesc=desc, itemref=item, qty=qty)
                db.session.add(data)
                db.session.commit()
        except BaseException as e:
            print("exception:-", e.args)
        return render_template('app1.html', message='submit')
    return render_template('app1.html', message='')


@app.route('/json')
def to_json():
    '''To send all data into json format'''
    data = pd.read_excel('Example_file.xlsx', sheet_name="BOM w Example")
    json_str = data.to_json(orient='records')
    return render_template('app1.html', message=json_str)


if __name__ == '__main__':
    app.run(debug=True)
